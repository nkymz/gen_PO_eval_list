# -*- coding: utf-8 -*-

import os
import re
import time

import openpyxl
import requests
from bs4 import BeautifulSoup

HTML_HEAD = """
<head>
<link rel="stylesheet" type="text/css" href="style.css">
<meta name="viewport" content="width=device-width,initial-scale=1">
</head>
"""

HTML_HEAD_NEW = """
<head>
<style>
table, th, td {
  border-collapse: collapse;
  border: 2px solid #ccc;
  line-height: 1.5;
}
th,td {
  padding: 4px;
}
th {
  background-color: #888888;
}
td.center {
  text-align: center;
}
</style>
</head>
"""


def gen_poh_eval_list(HDN_eval_list, uma_eval_list):
    DICT_EVAL = {"Ｓhyouka_siba":"5T", "Ａhyouka_siba":"4T", "Ｂhyouka_siba":"3T", "Ｃhyouka_siba":"2T",
                 "Ｄhyouka_siba":"1T", "Ｅhyouka_siba":"0T", "Ｓhyouka_dirt":"5D", "Ａhyouka_dirt":"4D",
                 "Ｂhyouka_dirt":"3D", "Ｃhyouka_dirt":"2D", "Ｄhyouka_dirt":"1D", "Ｅhyouka_dirtsiba":"0D"}
    path = os.getenv("HOMEDRIVE", "None") + os.getenv("HOMEPATH", "None") + "/Dropbox/POG/"
    wbpath = (path + "PO_HorseEvalList.xlsx").replace("\\", "/")
    wbbupath = (path + "PO_HorseEvalList_old.xlsx").replace("\\", "/")
    wb = openpyxl.load_workbook(wbpath)
    wb.save(wbbupath)
    ws = wb["POHEvalList"]
    xlrow = 1

    while ws.cell(row=xlrow, column=1).value:
        horse_name = ws.cell(row=xlrow, column=1).value

        if ws.cell(row=xlrow, column=6).value == "HDN_eval_new":
            ws.cell(row=xlrow, column=6).value = "HDN_eval_exist"
            xlrow += 1
            continue
        elif ws.cell(row=xlrow, column=6).value == "HDN_eval_exist":
            xlrow += 1
            continue

        for row in HDN_eval_list:
            if horse_name != row[4]:
                continue
            score = 0
            for i in range(3):
                myeval = DICT_EVAL[row[i + 10] + row[i + 13]]
                ws.cell(row=xlrow,column=7+i).value = myeval
                score += int(myeval[0]) * 7
            ws.cell(row=xlrow, column=10).value = score
            uma_score = ws.cell(row=xlrow, column=13).value
            ws.cell(row=xlrow, column=14).value = score * 2 if uma_score == "" else score + uma_score
            ws.cell(row=xlrow, column=6).value = "HDN_eval_new"
            break

        xlrow += 1

    xlrow = 1
    while ws.cell(row=xlrow, column=1).value:
        horse_name = ws.cell(row=xlrow, column=1).value
        if ws.cell(row=xlrow, column=11).value == "UMA_eval_new":
            ws.cell(row=xlrow, column=11).value = "UMA_eval_exist"
            xlrow += 1
            continue
        elif ws.cell(row=xlrow, column=11).value == "UMA_eval_exist":
            xlrow += 1
            continue
        for row in uma_eval_list:
            if horse_name != row[0]:
                continue
            ws.cell(row=xlrow, column=12).value = row[1]
            ws.cell(row=xlrow, column=13).value = row[2]
            hdn_score = ws.cell(row=xlrow, column=10).value
            ws.cell(row=xlrow, column=14).value = row[2] * 2 if hdn_score == "" else row[2] + hdn_score
            ws.cell(row=xlrow, column=11).value = "UMA_eval_new"
            break
        xlrow += 1

    wb.save(wbpath)

    xPOH_eval_list_all = [[cell.value for cell in row] for row in ws["A1:N" + str(xlrow - 1)]]
    return [row for row in xPOH_eval_list_all if row[5] != "HDN_eval_none" or row[10] != "UMA_eval_none"]


def gen_uma_eval_list():

    uma_url_1st_half = "http://umakeiba.com/post/category/%E5%84%AA%E9%A6%AC2%E6%AD%B3%E9%A6%AC%E3%83%81%E3%82%A7%E3%83%83%E3%82%AF/page/"

    uma_eval_list = []

    i = 1
    while True:
        target_url = uma_url_1st_half + str(i) + "/"

        time.sleep(1)
        r = requests.get(target_url)
        if r.status_code != requests.codes.ok:
            break
        soup = BeautifulSoup(r.content, 'lxml')

        for h2tag in soup.find_all("h2"):
            if not h2tag.find("a"):
                continue
            atag = h2tag.find("a")
            if not re.search(r"★評価一覧", atag.string):
                continue
            else:
                eval_page_url = atag.get("href")
                eval_page_no = int(eval_page_url.split("/")[-2])

            if eval_page_no < 6665:
                break
            time.sleep(1)
            r = requests.get(eval_page_url)
            if r.status_code != requests.codes.ok:
                break
            soup = BeautifulSoup(r.content, 'lxml')

            if len(soup.find_all("strong")) < 2:
                continue

            if soup.find_all("strong")[1].string[-1] == "点":
                eval_horses = [strong_tag.string for k, strong_tag in enumerate(soup.find_all("strong"))
                              if k % 2 == 0]
                eval_stars = [int(strong_tag.string[-2]) for k, strong_tag in enumerate(soup.find_all("strong"))
                              if k % 2 == 1]
            else:
                eval_horses = [strong_tag.string for k, strong_tag in enumerate(soup.find_all("strong"))
                               if k % 3 == 0]
                eval_stars = [int(strong_tag.string[-2]) for k, strong_tag in enumerate(soup.find_all("strong"))
                              if k % 3 == 2]
            for horse, star in zip(eval_horses, eval_stars):
                score = (star - 2 if star >= 2 else 0) * 5 * 3
                uma_eval_list.append([horse, star, score])
        else:
            i += 1
            continue
        break

    return uma_eval_list


def gen_HDN_eval_list():

    HDN_URL_1ST_HALF = "http://www.nikkankeiba.com/pog2018/hyouka/hyouka"
    url_2nd_half = ["01.html", "02.html", "03.html", "04.html", "05.html", "06.html", "07.html", "08.html,"
                    "09.html", "10.html", "11.html", "12.html"]
    HDN_eval_list = []

    for s in url_2nd_half:

        target_url = HDN_URL_1ST_HALF + s

        time.sleep(1)
        r = requests.get(target_url)
        if r.status_code != requests.codes.ok:
            break
        r.encoding = "euc-jp"
        soup = BeautifulSoup(r.text, 'lxml')

        mytable = soup.find("table").find_all("tr")

        for i, myrow in enumerate(mytable):
            if i == 0:
                continue
            eval_row = [element.string for element in myrow.find_all("td")]
            for j in range(3):
                eval_row.append(myrow.find_all("td")[j + 10].get("id"))

            HDN_eval_list.append(eval_row)

    return HDN_eval_list


def wrap_trtd(columns, tag_type):
    begin_tag_1sthalf = "<" + tag_type
    end_tag = "</" + tag_type + ">"
    t = "<tr>"
    for i, s in enumerate(columns):
        if i>= 2:
            t = t + begin_tag_1sthalf + ' class="center">' + s + end_tag
        else:
            t = t + begin_tag_1sthalf + '>' + s + end_tag
    t += "</tr>\n"

    return t


def deco_horse(string, status_hdn, status_uma, seal):
    string = string if seal == "-" else "<s>" + string + "</s>"
    if status_hdn == "HDN_eval_new" or status_uma == "UMA_eval_new":
        string = '<span style="font-weight: 900; color:#FF0000;">' + string + ' new!</span>'

    return string


def deco_hdn_eval(hdn_eval):
    if hdn_eval[1] == "T":
        if hdn_eval[0] == "5":
            s = '<span style="font-weight: 900; color:#FF0000;">' + hdn_eval[0] + '</span>'
        elif hdn_eval[0] == "4":
            s = '<span style="font-weight: 900; color:#0000FF;">' + hdn_eval[0] + '</span>'
        elif hdn_eval[0] == "3":
            s = '<span style="font-weight: 900;">' + hdn_eval[0] + '</span>'
        else:
            s = hdn_eval[0]
    else:
        if hdn_eval[0] == "5":
            s = '<span style="font-weight: 900; color:#000000; background-color:#FF0000">' + hdn_eval[0] + '</span>'
        elif hdn_eval[0] == "4":
            s = '<span style="font-weight: 900; color:#000000; background-color:#0000FF;">' + hdn_eval[0] + '</span>'
        elif hdn_eval[0] == "3":
            s = '<span style="font-weight: 900; color:#FFFFFF; background-color:#000000;">' + hdn_eval[0] + '</span>'
        else:
            s = '<span style="color:#FFFFFF; background-color:#000000;">' + hdn_eval[0] + '</span>'

    return s


def deco_uma_eval(uma_eval):
    if uma_eval in "9":
        s = '<span style="font-weight: 900; color:#FF0000;">' + uma_eval + '</span>'
    elif uma_eval in "78":
        s = '<span style="font-weight: 900; color:#0000FF;">' + uma_eval + '</span>'
    elif uma_eval in "6":
        s = '<span style="font-weight: 900;">' + uma_eval + '</span>'
    else:
        s = uma_eval

    return s


def out_poh_eval_list(xPOH_eval_list):
    path = os.getenv("HOMEDRIVE", "None") + os.getenv("HOMEPATH", "None") + "/Dropbox/POG/ppro_eval_list/"
    htmlpath = path + "index.html"
    f = open(htmlpath, mode="w", encoding="utf-8")
    f.write(HTML_HEAD)
    f.write('<table>\n')
    f.write(wrap_trtd(["馬名", "オーナー", "順位", "H1", "H2", "H3", "UM"], "th"))
    for row in xPOH_eval_list:
        horse_name = deco_horse(row[0], row[5], row[10], row[4])
        hdn_eval = []
        for i in range(3):
            hdn_eval.append(deco_hdn_eval(row[i + 6]))
        uma_eval = deco_uma_eval(str(row[11]))
        f.write(wrap_trtd([horse_name, row[1], row[2] + row[3], hdn_eval[0], hdn_eval[1], hdn_eval[2], uma_eval], "td"))
    f.write("\n")

    f.close()


def out_poh_eval_list_new(poh_eval_list_new):
    path = os.getenv("HOMEDRIVE", "None") + os.getenv("HOMEPATH", "None") + "/Dropbox/POG/"
    htmlpath = path + "ppro_eval_list_new.html"
    f = open(htmlpath, mode="w", encoding="utf-8")
    f.write(HTML_HEAD_NEW)
    f.write('<p> 全馬リストは<a href="https://nkymz.github.io/ppro_eval_list/">こちら</a>')
    f.write('<table>\n')
    f.write(wrap_trtd(["馬名", "オーナー", "順位", "H1", "H2", "H3", "UM"], "th"))
    for row in poh_eval_list_new:
        horse_name = deco_horse(row[0], "", "", row[4])
        hdn_eval = []
        for i in range(3):
            hdn_eval.append(deco_hdn_eval(row[i + 6]))
        uma_eval = deco_uma_eval(str(row[11]))
        f.write(wrap_trtd([horse_name, row[1], row[2] + row[3], hdn_eval[0], hdn_eval[1], hdn_eval[2], uma_eval], "td"))
    f.write("\n")

    f.close()


if __name__ == "__main__":
    xPOH_eval_list = gen_poh_eval_list(gen_HDN_eval_list(), gen_uma_eval_list())
    xPOH_eval_list.sort(key=lambda x:x[13], reverse=True)
    out_poh_eval_list(xPOH_eval_list)
    poh_eval_list_new = [r for r in xPOH_eval_list if r[5] == "HDN_eval_new" or r[10] == "UMA_eval_new"]
    out_poh_eval_list_new(poh_eval_list_new)





